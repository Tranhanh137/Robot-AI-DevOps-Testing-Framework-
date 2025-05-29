from openpyxl import load_workbook
from robot.api.deco import keyword
from robot.libraries.BuiltIn import BuiltIn
from faker import Faker
import allure

# Kh·ªüi t·∫°o Faker
fake = Faker('en_US')  # Ho·∫∑c fake = Faker('vi_VN') cho d·ªØ li·ªáu ti·∫øng Vi·ªát

# Dictionary √°nh x·∫° l·ªánh faker v·ªõi h√†m t∆∞∆°ng ·ª©ng
FAKER_MAPPINGS = {
    "{faker_name}": fake.name,
    "{faker_email}": fake.email,
    "{faker_phone}": fake.phone_number,
    "{faker_address}": fake.address,
    "{faker_text}": lambda: fake.text(max_nb_chars=50),
    "{faker_number}": lambda: fake.random_int(min=1, max=1000),
}

# C√°c action kh√¥ng c·∫ßn value (ch·∫°y n·∫øu value == "TRUE")
NO_VALUE_ACTIONS = {"clickButton", "checkCheckbox", "scrollToElement"}

def normalize_locator(raw_locator: str) -> str:
    """T·ª± ƒë·ªông th√™m prefix 'xpath=' n·∫øu c·∫ßn."""
    raw_locator = raw_locator.strip()
    if "=" in raw_locator.split()[0]:
        return raw_locator
    if raw_locator.startswith(("//", ".//", "(")):
        return f"xpath={raw_locator}"
    return raw_locator

def process_faker_value(value_str: str) -> str:
    """X·ª≠ l√Ω gi√° tr·ªã {faker_...} ƒë·ªÉ sinh d·ªØ li·ªáu ƒë·ªông."""
    value_str = value_str.strip()
    if value_str in FAKER_MAPPINGS:
        return str(FAKER_MAPPINGS[value_str]())
    return value_str

@keyword("Run Horizontal Test From Excel")
def run_horizontal_test_from_excel(file_path, test_case_name, sheet_name="Sheet1"):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    header = [cell.value for cell in ws[1]]
    try:
        tc_index = header.index(test_case_name)
    except ValueError:
        raise Exception(f"‚ùå Kh√¥ng t√¨m th·∫•y test case '{test_case_name}' trong file Excel.")

    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        action = row[0]
        locator = row[1]
        value = row[tc_index] if tc_index < len(row) else ""

        if not action:
            continue

        with allure.step(f"[Row {row_index}] {action} | locator={locator} | value={value}"):
            action = action.strip()
            locator = str(locator).strip() if locator else ""
            value_str = str(value).strip() if value is not None else ""
            normalized_value = value_str.upper()

            is_no_value_action = action in NO_VALUE_ACTIONS

            # üü° N·∫øu l√† action kh√¥ng c·∫ßn value ‚Üí ch·ªâ ch·∫°y khi value == TRUE
            if is_no_value_action:
                if normalized_value != "TRUE":
                    continue
                if locator:
                    locator = normalize_locator(locator)
                    locator = BuiltIn().replace_variables(locator)
                    BuiltIn().log(f"‚ñ∂Ô∏è [Row {row_index}] Action: {action} | Locator: {locator}", level="INFO")
                    try:
                        BuiltIn().run_keyword(action, locator)
                    except Exception as e:
                        BuiltIn().log(f"‚ùå [Row {row_index}] Failed: {action} with locator {locator}: {str(e)}", level="ERROR")
                        raise
                continue

            # üü¢ Ki·ªÉm tra n·∫øu kh√¥ng c√≥ locator v√† kh√¥ng c√≥ value ‚Üí b·ªè qua
            if not locator and not value_str:
                BuiltIn().log(f"‚è© [Row {row_index}] Skipped: No locator and no value for action {action}", level="INFO")
                continue

            # üü¢ V·ªõi action enterText, b·ªè qua n·∫øu value tr·ªëng
            if action == "enterText" and not value_str:
                BuiltIn().log(f"‚è© [Row {row_index}] Skipped: No value for action enterText", level="INFO")
                continue

            args = []
            if locator:
                locator = normalize_locator(locator)
                locator = BuiltIn().replace_variables(locator)
                args.append(locator)

            if action == "enterText":  # enterText lu√¥n c·∫ßn value
                # X·ª≠ l√Ω gi√° tr·ªã {faker_...} tr∆∞·ªõc khi thay bi·∫øn
                value_str = process_faker_value(value_str)
                value_str = BuiltIn().replace_variables(value_str)
                args.append(value_str)
            elif value_str:  # V·ªõi c√°c action kh√°c, ch·ªâ th√™m value n·∫øu kh√¥ng tr·ªëng
                value_str = process_faker_value(value_str)
                value_str = BuiltIn().replace_variables(value_str)
                args.append(value_str)

            BuiltIn().log(f"‚ñ∂Ô∏è [Row {row_index}] Action: {action} | Args: {args}", level="INFO")
            try:
                BuiltIn().run_keyword(action, *args)
            except Exception as e:
                BuiltIn().log(f"‚ùå [Row {row_index}] Failed: {action} with args {args}: {str(e)}", level="ERROR")
                raise